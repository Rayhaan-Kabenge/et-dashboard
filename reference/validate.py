import re
import sys, math
sys.path.insert(0, '/home/claude')
from openpyxl import load_workbook
from et_engine import Config, Stage, run

XLSM = "2017_2023_ETschudelr_Code.xlsm"   # place the workbook alongside this script

def na(v):
    if v is None: return None
    if isinstance(v, str) and v.startswith('#'): return None
    return v

def load_year(yr):
    wv = load_workbook(XLSM, data_only=True)
    wf = load_workbook(XLSM, keep_vba=True)
    ws = wv['Daily_%d' % yr]; ins = wv['Inputs']; wsf = wf['Daily_%d' % yr]
    m = re.search(r"MATCH\(\$F\d+,\$AF\$(\d+):\$AF\$(\d+),0\)", str(wsf['AC100'].value))
    sched_lo, sched_hi = (int(m.group(1)), int(m.group(2))) if m else (6, 31)
    cfg = Config(lat=ws['C9'].value, elev=ws['C10'].value, wndht=ws['C12'].value,
                 cn=ws['C20'].value, irrig_depth=ws['C25'].value, fert_depth=ws['C26'].value,
                 tall=(ws['C11'].value == 1),
                 dr0=(na(ws['D17'].value) or 0.0), dp0=(na(ws['C17'].value) or 0.0))
    soil = []
    for r in range(15, 25):
        top, bot, awc = ins['F%d'%r].value, ins['G%d'%r].value, ins['H%d'%r].value
        if top is not None: soil.append((top, bot, awc))
    stages = []
    for r in range(5, 30):
        d = ws['I%d'%r].value
        if d is None: continue
        stages.append(Stage(date=d.date(), label=ws['J%d'%r].value,
                            slope=na(ws['S%d'%r].value) or 0.0, intercept=na(ws['T%d'%r].value) or 0.0,
                            managed_depth=na(ws['U%d'%r].value), mad=na(ws['V%d'%r].value)))
    schedule = {}
    for r in range(sched_lo, sched_hi+1):
        d = ws['AF%d'%r].value
        if d is not None: schedule[d.date()] = ws['AE%d'%r].value
    weather, gold = [], []
    for r in range(35, 215):
        fv=ws['F%d'%r].value
        if not isinstance(ws['N%d'%r].value,(int,float)) or not hasattr(fv,'date'): continue
        weather.append(dict(date=ws['F%d'%r].value.date(), doy=ws['H%d'%r].value,
                            tmax=ws['N%d'%r].value, tmin=ws['O%d'%r].value, ea=ws['P%d'%r].value,
                            rs=ws['Q%d'%r].value, u=ws['R%d'%r].value, precip=ws['S%d'%r].value))
        gold.append(dict(gdd=ws['J%d'%r].value, cumgdd=ws['K%d'%r].value, stage=ws['L%d'%r].value or '',
                        interval=ws['M%d'%r].value, ro=ws['T%d'%r].value, etr=ws['U%d'%r].value,
                        fracint=na(ws['V%d'%r].value), kcr=na(ws['W%d'%r].value), etc=na(ws['X%d'%r].value),
                        dp=na(ws['Y%d'%r].value), depletion=na(ws['Z%d'%r].value), ad=na(ws['AA%d'%r].value),
                        should=ws['AB%d'%r].value, applied=ws['AC%d'%r].value))
    return cfg, soil, stages, schedule, weather, gold

def cmp_year(yr):
    cfg, soil, stages, schedule, weather, gold = load_year(yr)
    etr_inj = [g['etr'] for g in gold]   # inject Excel ETr -> tests B-I logic exactly
    out = run(cfg, soil, stages, schedule, weather, etr_override=etr_inj)
    numcols = ['gdd','cumgdd','ro','fracint','kcr','etc','dp','depletion','ad']
    maxd = {c: 0.0 for c in numcols}
    badbool = badstage = badint = 0
    worst = {}
    for o, g in zip(out, gold):
        for c in numcols:
            a, b = o[c], g[c]
            if a is None or b is None:
                if not (a is None and b is None): maxd[c] = max(maxd[c], 9e9)
                continue
            d = abs(a-b)
            if d > maxd[c]: maxd[c] = d; worst[c] = (yr, o['date'], a, b)
        if bool(o['should_irrigate']) != bool(g['should']): badbool += 1
        if (o['stage'] or '') != (g['stage'] or ''): badstage += 1
        if o['interval'] != g['interval']: badint += 1
    appd = max(abs((o['applied'] or 0)-(g['applied'] or 0)) for o,g in zip(out,gold))
    print(f'--- {yr}  (n={len(out)} days) ---')
    for c in numcols: print(f'   {c:10} max|diff| = {maxd[c]:.2e}')
    print(f'   applied    max|diff| = {appd:.2e}')
    print(f'   stage mismatches={badstage}  interval mismatches={badint}  should-irrigate mismatches={badbool}')
    return maxd, worst

if __name__ == '__main__':
    allworst = {}
    for yr in [2017,2018,2019,2020,2021,2022,2023]:
        md, w = cmp_year(yr); allworst.update(w)
